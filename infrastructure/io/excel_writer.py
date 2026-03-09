from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.settings import ALERTS_DIR

logger = logging.getLogger(__name__)

STATUS_FILLS: dict[str, str] = {
    "CONFORME": "C6EFCE",
    "PARCIAL": "FFEB9C",
    "NÃO CONFORME": "FFC7CE",
    "INCOMPLETE": "FFCC99",
    "FAILED": "D9D9D9",
}
HEADER_FILL = "1F3864"
HEADER_FONT_COLOR = "FFFFFF"

_S1_COLS = ["Métrica", "Valor"]

_S2_COLS = [
    "Processo ID",
    "Empresa",
    "Valor Contrato",
    "Status",
    "Score",
    "Nível Diagnóstico",
    "R001",
    "R002",
    "R003",
    "R004",
    "Dias p/ Publicar",
    "Violação Principal",
    "Severidade",
    "Requer Revisão",
    "Flags",
]

_S6_COLS = ["Parâmetro", "Valor"]


def _sanitize(pid: str) -> str:
    try:
        return str(pid).replace("/", "_").replace("\\", "_")
    except Exception as exc:
        logger.warning("Failed to sanitize pid '%s': %s", pid, exc)
        return ""


def _safe_str(value: Any) -> str:
    try:
        if value is None:
            return ""
        return str(value)
    except Exception:
        return ""


def _safe_float(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def _safe_bool(value: Any) -> bool:
    try:
        return bool(value)
    except Exception:
        return False


def _safe_list(value: Any) -> list:
    try:
        return value if isinstance(value, list) else []
    except Exception:
        return []


def _make_header_row(ws, cols: list[str]) -> None:
    try:
        for col_index, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_index, value=col_name)
            cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
            cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    except Exception as exc:
        logger.warning("Failed to create header row on sheet '%s': %s", _safe_str(ws.title), exc)


def _apply_status_fill(cell, status: str) -> None:
    try:
        fill_hex = STATUS_FILLS.get(_safe_str(status), "")
        if fill_hex:
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    except Exception as exc:
        logger.warning("Failed applying status fill for status '%s': %s", status, exc)


def _set_column_widths(ws, widths: dict[int, int]) -> None:
    try:
        for col_index, width in widths.items():
            ws.column_dimensions[get_column_letter(int(col_index))].width = int(width)
    except Exception as exc:
        logger.warning("Failed setting column widths on sheet '%s': %s", _safe_str(ws.title), exc)


def _freeze_and_filter(ws) -> None:
    try:
        ws.freeze_panes = ws["A2"]
        ws.auto_filter.ref = ws.dimensions
    except Exception as exc:
        logger.warning("Failed applying freeze/filter on sheet '%s': %s", _safe_str(ws.title), exc)


def _build_sheet1_resumo(ws, aggregate: dict, analyst_name: str, software_version: str) -> None:
    try:
        ws.title = "Resumo Executivo"
        _make_header_row(ws, _S1_COLS)

        coverage = aggregate.get("coverage", {}) if isinstance(aggregate, dict) else {}
        conformity = aggregate.get("conformity_summary", {}) if isinstance(aggregate, dict) else {}
        rule_averages = aggregate.get("rule_averages", {}) if isinstance(aggregate, dict) else {}

        rows = [
            ("Relatório gerado em", _safe_str(aggregate.get("generated_at", ""))),
            ("Analista", _safe_str(analyst_name) or "Sistema"),
            ("Versão", _safe_str(software_version)),
            ("Contratos Descobertos", int(coverage.get("total_discovered", 0) or 0)),
            ("Contratos Extraídos", int(coverage.get("total_extracted", 0) or 0)),
            ("Publicações Encontradas", int(coverage.get("total_pub_found", 0) or 0)),
            ("Contratos Analisados", int(coverage.get("total_analyzed", 0) or 0)),
            ("Taxa de Cobertura", f"{_safe_float(coverage.get('coverage_rate', 0.0)):.1%}"),
            (
                "Taxa de Conformidade",
                f"{_safe_float(conformity.get('overall_conformity_rate', 0.0)):.1%}",
            ),
            ("Score Médio", f"{_safe_float(conformity.get('average_score', 0.0)):.1f}"),
            ("CONFORME", int(conformity.get("CONFORME", 0) or 0)),
            ("PARCIAL", int(conformity.get("PARCIAL", 0) or 0)),
            ("NÃO CONFORME", int(conformity.get("NÃO CONFORME", 0) or 0)),
            ("INCOMPLETE", int(conformity.get("INCOMPLETE", 0) or 0)),
            ("Média R001", f"{_safe_float(rule_averages.get('R001', 0.0)):.1f}"),
            ("Média R002", f"{_safe_float(rule_averages.get('R002', 0.0)):.1f}"),
            ("Média R003", f"{_safe_float(rule_averages.get('R003', 0.0)):.1f}"),
            ("Média R004", f"{_safe_float(rule_averages.get('R004', 0.0)):.1f}"),
        ]

        for row_index, (metric, value) in enumerate(rows, start=2):
            ws.cell(row=row_index, column=1, value=metric)
            ws.cell(row=row_index, column=2, value=value)

        _set_column_widths(ws, {1: 40, 2: 28})
    except Exception as exc:
        logger.warning("Failed building sheet 1: %s", exc)


def _contract_to_s2_row(contract: dict) -> list[Any]:
    try:
        flags = _safe_list(contract.get("flags", [])) if isinstance(contract, dict) else []
        return [
            _safe_str(contract.get("processo_id", "")),
            _safe_str(contract.get("company_name", "")),
            _safe_str(contract.get("contract_value", "")),
            _safe_str(contract.get("overall_status", "")),
            _safe_float(contract.get("conformity_score", 0.0)),
            _safe_str(contract.get("agreement_level", "")),
            _safe_str(contract.get("R001_verdict", "")),
            _safe_str(contract.get("R002_verdict", "")),
            _safe_str(contract.get("R003_verdict", "")),
            _safe_str(contract.get("R004_verdict", "")),
            contract.get("days_to_publish", None),
            _safe_str(contract.get("primary_violation", "")),
            _safe_str(contract.get("severity", "")),
            _safe_bool(contract.get("requires_review", False)),
            ", ".join(_safe_str(item) for item in flags),
        ]
    except Exception as exc:
        logger.warning("Failed mapping contract row: %s", exc)
        return ["", "", "", "", 0.0, "", "", "", "", "", None, "", "", False, ""]


def _build_sheet2_resultados(ws, contracts: list[dict]) -> None:
    try:
        ws.title = "Resultados Detalhados"
        _make_header_row(ws, _S2_COLS)

        rows = contracts if isinstance(contracts, list) else []
        for row_index, contract in enumerate(rows, start=2):
            data_row = _contract_to_s2_row(contract if isinstance(contract, dict) else {})
            for col_index, value in enumerate(data_row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)
            _apply_status_fill(ws.cell(row=row_index, column=4), _safe_str(data_row[3]))

        widths = {index: 12 for index in range(1, len(_S2_COLS) + 1)}
        widths[1] = 25
        widths[2] = 30
        widths[4] = 15
        widths[5] = 8
        _set_column_widths(ws, widths)
        _freeze_and_filter(ws)
    except Exception as exc:
        logger.warning("Failed building sheet 2: %s", exc)


def _load_alert_level_for_contract(contract: dict) -> str:
    try:
        if not isinstance(contract, dict):
            return ""
        pid_safe = _safe_str(contract.get("pid_safe", ""))
        if not pid_safe:
            pid_safe = _sanitize(_safe_str(contract.get("processo_id", "")))
        if not pid_safe:
            return ""

        candidates = [
            ALERTS_DIR / f"{pid_safe}_alerts.json",
            ALERTS_DIR / f"{pid_safe}_alert.json",
        ]
        for path in candidates:
            if not path.exists():
                continue
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return _safe_str(data.get("alert_level", ""))
        return ""
    except Exception as exc:
        logger.warning("Failed loading alert level: %s", exc)
        return ""


def _build_sheet3_nao_conformes(ws, contracts: list[dict]) -> None:
    try:
        ws.title = "Não Conformes"
        rows = contracts if isinstance(contracts, list) else []

        filtered: list[dict] = []
        for contract in rows:
            if not isinstance(contract, dict):
                continue
            status = _safe_str(contract.get("overall_status", ""))
            alert_level = _load_alert_level_for_contract(contract)
            if status in ("NÃO CONFORME", "PARCIAL") or alert_level == "FAILED":
                filtered.append(contract)

        if not filtered:
            ws["A1"] = "Nenhum contrato não conforme encontrado."
            return

        _make_header_row(ws, _S2_COLS)
        for row_index, contract in enumerate(filtered, start=2):
            data_row = _contract_to_s2_row(contract)
            for col_index, value in enumerate(data_row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)
            _apply_status_fill(ws.cell(row=row_index, column=4), _safe_str(data_row[3]))

        widths = {index: 12 for index in range(1, len(_S2_COLS) + 1)}
        widths[1] = 25
        widths[2] = 30
        widths[4] = 15
        widths[5] = 8
        _set_column_widths(ws, widths)
        _freeze_and_filter(ws)
    except Exception as exc:
        logger.warning("Failed building sheet 3: %s", exc)


def _missing_phase(pipeline_stage: str) -> str:
    try:
        mapping = {
            "DISCOVERED": "Extração de Contrato",
            "EXTRACTED": "Extração de Publicação",
            "PUB_FOUND": "Pré-processamento",
            "PREPROCESSED": "Análise de Conformidade",
        }
        return _safe_str(mapping.get(_safe_str(pipeline_stage), ""))
    except Exception:
        return ""


def _build_sheet4_incompletos(ws, contracts: list[dict]) -> None:
    try:
        ws.title = "Incompletos e Publicações Ausentes"
        cols = ["Processo ID", "Empresa", "Etapa Pipeline", "Fase Faltante"]
        rows = contracts if isinstance(contracts, list) else []
        filtered = [
            item
            for item in rows
            if isinstance(item, dict)
            and _safe_str(item.get("pipeline_stage", "")) not in ("SCORED", "COMPLIANCE")
        ]

        _make_header_row(ws, cols)
        for row_index, contract in enumerate(filtered, start=2):
            pipeline_stage = _safe_str(contract.get("pipeline_stage", ""))
            ws.cell(row=row_index, column=1, value=_safe_str(contract.get("processo_id", "")))
            ws.cell(row=row_index, column=2, value=_safe_str(contract.get("company_name", "")))
            ws.cell(row=row_index, column=3, value=pipeline_stage)
            ws.cell(row=row_index, column=4, value=_missing_phase(pipeline_stage))

        _set_column_widths(ws, {1: 25, 2: 30, 3: 20, 4: 28})
        _freeze_and_filter(ws)
    except Exception as exc:
        logger.warning("Failed building sheet 4: %s", exc)


def _build_sheet5_erros(ws, contracts: list[dict]) -> None:
    try:
        ws.title = "Erros de Extração"
        rows = contracts if isinstance(contracts, list) else []
        filtered = [
            item
            for item in rows
            if isinstance(item, dict) and _safe_bool(item.get("error_flag", False))
        ]

        if not filtered:
            ws["A1"] = "Nenhum erro de extração registrado."
            return

        cols = ["Processo ID", "Empresa", "Etapa Pipeline", "Flag de Erro"]
        _make_header_row(ws, cols)
        for row_index, contract in enumerate(filtered, start=2):
            ws.cell(row=row_index, column=1, value=_safe_str(contract.get("processo_id", "")))
            ws.cell(row=row_index, column=2, value=_safe_str(contract.get("company_name", "")))
            ws.cell(row=row_index, column=3, value=_safe_str(contract.get("pipeline_stage", "")))
            ws.cell(row=row_index, column=4, value=True)

        _set_column_widths(ws, {1: 25, 2: 30, 3: 20, 4: 16})
        _freeze_and_filter(ws)
    except Exception as exc:
        logger.warning("Failed building sheet 5: %s", exc)


def _build_sheet6_metadados(ws, aggregate: dict, analyst_name: str) -> None:
    try:
        ws.title = "Metadados do Relatório"
        _make_header_row(ws, _S6_COLS)

        coverage = aggregate.get("coverage", {}) if isinstance(aggregate, dict) else {}
        rows = [
            ("Sistema", "TCM-RJ Análise de Contratos"),
            ("Analista", _safe_str(analyst_name) or "Sistema"),
            ("Data de Geração", _safe_str(aggregate.get("generated_at", ""))),
            ("Total de Contratos", int(coverage.get("total_discovered", 0) or 0)),
            ("Contratos Analisados", int(coverage.get("total_analyzed", 0) or 0)),
        ]

        for row_index, (param, value) in enumerate(rows, start=2):
            ws.cell(row=row_index, column=1, value=param)
            ws.cell(row=row_index, column=2, value=value)

        _set_column_widths(ws, {1: 34, 2: 34})
    except Exception as exc:
        logger.warning("Failed building sheet 6: %s", exc)


def write_excel_report(
    aggregate: dict,
    output_path: Path,
    analyst_name: str = "",
    software_version: str = "1.0.0",
    year_filter: str = "",
) -> Path:
    try:
        if not isinstance(aggregate, dict) or not aggregate or "contracts" not in aggregate:
            raise ValueError("aggregate must be a non-empty dict containing 'contracts'")

        wb = openpyxl.Workbook()
        active_sheet = wb.active
        if active_sheet is not None:
            wb.remove(active_sheet)

        contracts = aggregate.get("contracts", []) if isinstance(aggregate.get("contracts", []), list) else []

        ws1 = wb.create_sheet("Resumo Executivo")
        _build_sheet1_resumo(ws1, aggregate, analyst_name, software_version)

        ws2 = wb.create_sheet("Resultados Detalhados")
        _build_sheet2_resultados(ws2, contracts)

        ws3 = wb.create_sheet("Não Conformes")
        _build_sheet3_nao_conformes(ws3, contracts)

        ws4 = wb.create_sheet("Incompletos e Publicações Ausentes")
        _build_sheet4_incompletos(ws4, contracts)

        ws5 = wb.create_sheet("Erros de Extração")
        _build_sheet5_erros(ws5, contracts)

        ws6 = wb.create_sheet("Metadados do Relatório")
        _build_sheet6_metadados(ws6, aggregate, analyst_name)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path
    except Exception as exc:
        logger.warning("Failed to write excel report: %s", exc)
        return output_path


def write_excel_filtered(contracts: list[dict], output_path: Path, analyst_name: str = "") -> Path:
    try:
        wb = openpyxl.Workbook()
        active_sheet = wb.active
        if active_sheet is not None:
            wb.remove(active_sheet)

        ws1 = wb.create_sheet("Resultados")
        _build_sheet2_resultados(ws1, contracts if isinstance(contracts, list) else [])

        ws2 = wb.create_sheet("Metadados")
        _make_header_row(ws2, _S6_COLS)
        rows = [
            ("Analista", _safe_str(analyst_name) or "Sistema"),
            ("Data de Geração", datetime.now().isoformat()),
        ]
        for row_index, (key, value) in enumerate(rows, start=2):
            ws2.cell(row=row_index, column=1, value=key)
            ws2.cell(row=row_index, column=2, value=value)
        _set_column_widths(ws2, {1: 24, 2: 36})

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path
    except Exception as exc:
        logger.warning("Failed to write filtered excel report: %s", exc)
        return output_path
