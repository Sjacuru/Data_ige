"""tests/test_stage7_dashboard.py

Stage 7 — Dashboard Infrastructure Test Suite
Offline only. No browser, no Streamlit server, no network.
"""
from __future__ import annotations

import importlib
import json
import shutil
import sys
import tempfile
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

# ── Patch st.cache_data before any dashboard import ──────────────────────────
import streamlit as st

st.cache_data = lambda **kw: (lambda f: f)

PASSED = 0
FAILED = 0
WARNINGS = 0

GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
BOLD = "\033[1m"
RESET = "\033[0m"


def section(title: str) -> None:
    print(f"\n{BOLD}── {title} ──{RESET}")


def check(label: str, condition: bool, hint: str = "") -> None:
    global PASSED, FAILED
    if condition:
        print(f"  {GREEN}✓{RESET} {label}")
        PASSED += 1
    else:
        msg = f"  {RED}✗{RESET} {label}"
        if hint:
            msg += f"  [{hint}]"
        print(msg)
        FAILED += 1


def warn(msg: str) -> None:
    global WARNINGS
    print(f"  {YELLOW}⚠{RESET} {msg}")
    WARNINGS += 1


def info(msg: str) -> None:
    print(f"    {msg}")


def _make_temp_data_dir() -> tuple[Path, dict]:
    tmp = Path(tempfile.mkdtemp())
    pids = ["TEST-ORG/00001", "TEST-ORG/00002", "TEST-ORG/00003"]
    pid_safes = [pid.replace("/", "_") for pid in pids]

    disc = tmp / "discovery"
    disc.mkdir(parents=True, exist_ok=True)
    (disc / "processo_links.json").write_text(
        json.dumps(
            {
                "discovery_date": "2026-03-09",
                "total_companies": 1,
                "total_processos": 3,
                "processos": [{"processo_id": pid, "url": "https://test"} for pid in pids],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    (tmp / "compliance_progress.json").write_text(
        json.dumps(
            {
                "last_run": "2026-03-09T10:00:00",
                "stats": {"total": 3, "completed": 2, "failed": 1, "skipped": 0},
                "completed": pids[:2],
                "failed": [{"processo_id": pids[2], "error": "Timeout", "at": "2026-03-09T10:00:00"}],
                "skipped": [],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    conf_dir = tmp / "conformity"
    conf_dir.mkdir(parents=True, exist_ok=True)
    (conf_dir / f"{pid_safes[0]}_conformity.json").write_text(
        json.dumps(
            {
                "processo_id": pids[0],
                "overall_status": "CONFORME",
                "conformity_score": 95.0,
                "flags": [],
                "requires_review": False,
                "score_breakdown": {
                    "R001": {"verdict": "PASS", "score": 40.0},
                    "R002": {"verdict": "PASS", "score": 30.0},
                    "R003": {"verdict": "PASS", "score": 20.0},
                    "R004": {"verdict": "PASS", "score": 10.0},
                },
                "diagnostic": {"agreement_level": "CONFIRMED"},
                "recommendations": [],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    alert_dir = tmp / "alerts"
    alert_dir.mkdir(parents=True, exist_ok=True)
    (alert_dir / f"{pid_safes[0]}_alert.json").write_text(
        json.dumps(
            {
                "processo_id": pids[0],
                "evaluated_at": "2026-03-09T10:00:00",
                "alert_level": "OK",
                "reason": "CONFORME",
                "reason_details": {
                    "action": "no_action",
                    "evidence": "conforme",
                    "failed_rules": [],
                    "recommendations": [],
                },
                "overall_status": "CONFORME",
                "conformity_score": 95.0,
                "flags": [],
                "failed_rules": [],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return tmp, {"pids": pids, "pid_safes": pid_safes}


def track_a_imports() -> None:
    section("TRACK A — Imports")

    try:
        import infrastructure.dashboard.pipeline_runner as _pr

        _ = _pr
        check("A1: pipeline_runner imports cleanly", True)
    except Exception as exc:
        check("A1: pipeline_runner imports cleanly", False, hint=str(exc))

    try:
        import infrastructure.dashboard.state_reader as _sr

        _ = _sr
        check("A2: state_reader imports cleanly", True)
    except Exception as exc:
        check("A2: state_reader imports cleanly", False, hint=str(exc))

    try:
        import infrastructure.dashboard.chart_builder as _cb

        _ = _cb
        check("A3: chart_builder imports cleanly", True)
    except Exception as exc:
        check("A3: chart_builder imports cleanly", False, hint=str(exc))

    try:
        from application.app import render_app

        check("A4: render_app imports without Streamlit side effects", callable(render_app))
    except Exception as exc:
        check("A4: render_app imports without Streamlit side effects", False, hint=str(exc))


def track_b_pipeline_runner() -> None:
    section("TRACK B — Pipeline Runner")

    from infrastructure.dashboard.pipeline_runner import (
        STAGE_PROGRESS_FILES,
        STAGE_SCRIPTS,
        _parse_stage_progress,
        get_stage_status,
        is_any_running,
        launch_stage,
    )

    import infrastructure.dashboard.pipeline_runner as pr

    pr._RUNNING_STAGE = "stage4"
    result = launch_stage("stage5")
    pr._RUNNING_STAGE = None
    check("B1: launch_stage blocked when already running", result is False)

    s = get_stage_status("stage4")
    check("B2: get_stage_status returns valid status", s["status"] in ("NOT_STARTED", "COMPLETE", "IN_PROGRESS", "FAILED"))
    check("B2: get_stage_status never raises", True)

    tmp = Path(tempfile.mkdtemp())
    prog_data = {
        "stats": {"total": 10, "completed": 7, "failed": 1, "skipped": 0},
        "completed": ["A"] * 7,
        "failed": [{"processo_id": "X", "error": "e", "at": "t"}],
        "last_run": "2026-01-01T00:00:00",
    }
    (tmp / "compliance_progress.json").write_text(json.dumps(prog_data, ensure_ascii=False, indent=2), encoding="utf-8")
    original_file = pr.STAGE_PROGRESS_FILES["stage4"]
    pr.STAGE_PROGRESS_FILES["stage4"] = tmp / "compliance_progress.json"
    s4 = get_stage_status("stage4")
    pr.STAGE_PROGRESS_FILES["stage4"] = original_file
    shutil.rmtree(tmp, ignore_errors=True)
    check("B3: stage4 total=10", s4["total"] == 10, hint=str(s4["total"]))
    check("B3: stage4 completed=7", s4["completed"] == 7, hint=str(s4["completed"]))
    check("B3: stage4 failed_count=1", s4["failed_count"] == 1)

    tmp2 = Path(tempfile.mkdtemp())
    (tmp2 / "conformity_summary.json").write_text(
        json.dumps({"total_contracts": 50, "generated_at": "2026-01-01T00:00:00"}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    original5 = pr.STAGE_PROGRESS_FILES["stage5"]
    pr.STAGE_PROGRESS_FILES["stage5"] = tmp2 / "conformity_summary.json"
    s5 = get_stage_status("stage5")
    pr.STAGE_PROGRESS_FILES["stage5"] = original5
    shutil.rmtree(tmp2, ignore_errors=True)
    check("B4: stage5 total=50 from summary schema", s5["total"] == 50, hint=str(s5["total"]))

    total, completed, _failed_count = _parse_stage_progress(
        "stage4",
        {"stats": {"total": 400, "completed": 240, "failed": 0}, "completed": ["x"] * 240, "failed": []},
    )
    pct = round((completed / total) * 100, 1) if total else 0.0
    check("B5: progress_pct 240/400 = 60.0", pct == 60.0, hint=str(pct))

    check("B6: is_any_running() False at rest", is_any_running() is False)

    check("B7: STAGE_PROGRESS_FILES has 7 keys", len(STAGE_PROGRESS_FILES) == 7, hint=str(list(STAGE_PROGRESS_FILES.keys())))

    check(
        "B8: STAGE_SCRIPTS['stage1'] == 'application/main.py'",
        STAGE_SCRIPTS["stage1"] == "application/main.py",
        hint=str(STAGE_SCRIPTS.get("stage1", "")),
    )


def track_c_state_reader() -> None:
    section("TRACK C — State Reader")

    import infrastructure.dashboard.state_reader as sr

    detail = sr.read_processo_detail("COMPLETELY_UNKNOWN_PID_9999")
    check(
        "C1: unknown PID returns all-None dict",
        all(value is None for value in detail.values()),
        hint=str({key: value for key, value in detail.items() if value}),
    )

    conf_dir = ROOT / "data" / "conformity"
    files = list(conf_dir.glob("*_conformity.json")) if conf_dir.exists() else []
    if files:
        pid_safe = files[0].stem.replace("_conformity", "")
        d = sr.read_processo_detail(pid_safe)
        check("C2: known PID conformity is not None", d["conformity"] is not None)
    else:
        warn("C2: skipped — no conformity files on disk")

    called: list[int] = []
    import domain.services.alert_queue as aq

    tmp_alert_root = Path(tempfile.mkdtemp())
    temp_alerts_dir = tmp_alert_root / "alerts"
    temp_alerts_dir.mkdir(parents=True, exist_ok=True)
    (temp_alerts_dir / "TEST_ORG_00001_alert.json").write_text(
        json.dumps(
            {
                "processo_id": "TEST-ORG/00001",
                "alert_level": "OK",
                "reason": "CONFORME",
                "conformity_score": 95.0,
                "flags": [],
                "failed_rules": [],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    orig_alerts_dir = sr.ALERTS_DIR
    orig_builder = aq.build_alert_queue
    try:
        sr.ALERTS_DIR = temp_alerts_dir
        aq.build_alert_queue = lambda alerts: (called.append(1), orig_builder(alerts))[1]
        sr.read_all_alerts()
    finally:
        aq.build_alert_queue = orig_builder
        sr.ALERTS_DIR = orig_alerts_dir
        shutil.rmtree(tmp_alert_root, ignore_errors=True)
    check("C3: read_all_alerts calls build_alert_queue", len(called) > 0)

    errors = sr.read_errors()
    check("C4: read_errors returns stage2/3/4 keys", {"stage2", "stage3", "stage4"} == set(errors.keys()))
    if (ROOT / "data" / "compliance_progress.json").exists():
        prog = json.loads((ROOT / "data" / "compliance_progress.json").read_text(encoding="utf-8"))
        expected_fails = len(prog.get("failed", []))
        check(
            "C4: stage4 error count matches compliance_progress.json",
            len(errors["stage4"]) == expected_fails,
            hint=f"expected {expected_fails}, got {len(errors['stage4'])}",
        )
    else:
        warn("C4: compliance_progress.json not found — partial check only")

    lines = sr.read_log_tail("completely_nonexistent_stage_xyz")
    check("C5: read_log_tail returns [] for missing log", lines == [], hint=str(lines))

    output_files = sr.list_output_files()
    check("C6: list_output_files returns list", isinstance(output_files, list))

    tmp_root = Path(tempfile.mkdtemp())
    orig_alerts = sr.ALERTS_DIR if hasattr(sr, "ALERTS_DIR") else None
    cfg = None
    try:
        import config.settings as cfg

        cfg.ALERTS_DIR = tmp_root / "alerts_empty"
        _sr2_detail = sr.read_processo_detail("ANY/PID")
        _ = _sr2_detail
        check("C7: read_processo_detail does not raise on empty dir", True)
    except Exception as exc:
        check("C7: read_processo_detail does not raise on empty dir", False, hint=str(exc))
    finally:
        if orig_alerts and cfg is not None:
            cfg.ALERTS_DIR = orig_alerts
        shutil.rmtree(tmp_root, ignore_errors=True)


def track_d_carryover() -> None:
    section("TRACK D — Carryover Smoke Tests (CT1–CT6)")

    tmp, pid_map = _make_temp_data_dir()
    _ = pid_map

    try:
        from infrastructure.io.state_index_builder import build_state_index
        import infrastructure.io.state_index_builder as sib

        orig_disc = sib.DISCOVERY_FILE
        orig_conf_dir = sib.CONFORMITY_DIR
        sib.DISCOVERY_FILE = tmp / "discovery" / "processo_links.json"
        sib.CONFORMITY_DIR = tmp / "conformity"
        idx = build_state_index(sib.DISCOVERY_FILE)
        sib.DISCOVERY_FILE = orig_disc
        sib.CONFORMITY_DIR = orig_conf_dir
        check("D1: build_state_index runs without error", isinstance(idx, dict))
        check("D1: total_pids == 3", idx.get("total_pids") == 3, hint=str(idx.get("total_pids")))

        from infrastructure.io.report_aggregator import build_aggregate_report
        import infrastructure.io.report_aggregator as ra

        ra_disc_orig = ra.DISCOVERY_FILE
        ra.DISCOVERY_FILE = tmp / "discovery" / "processo_links.json"
        agg = build_aggregate_report(state_index=idx)
        ra.DISCOVERY_FILE = ra_disc_orig
        check(
            "D2: build_aggregate_report total_discovered == 3",
            agg["coverage"]["total_discovered"] == 3,
            hint=str(agg["coverage"]["total_discovered"]),
        )

        import openpyxl
        from infrastructure.io.excel_writer import write_excel_report

        p_full = tmp / "full_report.xlsx"
        write_excel_report(agg, p_full)
        wb = openpyxl.load_workbook(p_full)
        check("D3: full Excel has 6 sheets", len(wb.sheetnames) == 6, hint=str(wb.sheetnames))

        from infrastructure.io.excel_writer import write_excel_filtered

        p_filt = tmp / "filtered.xlsx"
        write_excel_filtered(agg.get("contracts", [])[:2], p_filt)
        wb2 = openpyxl.load_workbook(p_filt)
        check("D4: filtered Excel has 2 sheets", len(wb2.sheetnames) == 2, hint=str(wb2.sheetnames))

        import csv
        import io as _io
        from infrastructure.io.report_csv_writer import REPORT_CSV_COLUMNS, write_report_csv

        p_csv = tmp / "report.csv"
        write_report_csv(agg.get("contracts", []), p_csv, "2026-03-09T00:00:00")
        with open(p_csv, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            cols = reader.fieldnames or []
        check("D5: CSV has 19 columns", len(cols) == 19, hint=str(len(cols)))
        check("D5: CSV columns match REPORT_CSV_COLUMNS", cols == REPORT_CSV_COLUMNS, hint=str(cols))

        import pandas as pd
        from infrastructure.io.report_csv_writer import build_report_csv_bytes

        b = build_report_csv_bytes(agg.get("contracts", []), "2026-03-09T00:00:00")
        df = pd.read_csv(_io.BytesIO(b), encoding="utf-8-sig")
        check("D6: build_report_csv_bytes parseable by pandas", len(df.columns) == 19, hint=str(len(df.columns)))

        from infrastructure.io.aggregate_json_writer import write_aggregate_json

        p_json = tmp / "aggregate.json"
        write_aggregate_json(agg, p_json)
        raw = p_json.read_text(encoding="utf-8")
        parsed = json.loads(raw)
        check("D7: aggregate JSON is valid", "coverage" in parsed)

        from application.workflows.stage6_report import run_stage6_report

        result = run_stage6_report(state_only=True)
        required_keys = {
            "status",
            "state_index_path",
            "excel_path",
            "csv_path",
            "json_path",
            "log_path",
            "total_contracts",
            "total_analyzed",
            "coverage_rate",
            "conformity_rate",
            "processing_time_seconds",
            "errors",
        }
        check(
            "D8: run_stage6_report has all 12 required keys",
            required_keys.issubset(result.keys()),
            hint=f"missing: {required_keys - result.keys()}",
        )

        import sys as _sys

        _before_mods = set(_sys.modules.keys())
        _ = _before_mods
        run_stage6_report(state_only=True)
        check(
            "D9: full_pipeline not in sys.modules after run_stage6_report",
            "application.workflows.full_pipeline" not in _sys.modules,
            hint="full_pipeline was imported as side effect",
        )

    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main() -> int:
    print(f"\n{BOLD}{'═' * 65}{RESET}")
    print(f"{BOLD}  STAGE 7 — DASHBOARD TEST SUITE{RESET}")
    print(f"{BOLD}{'═' * 65}{RESET}")

    track_a_imports()
    track_b_pipeline_runner()
    track_c_state_reader()
    track_d_carryover()

    print(f"\n{BOLD}{'═' * 65}{RESET}")
    print(f"{BOLD}  RESULTS{RESET}")
    print(f"{'═' * 65}")
    print(f"  {GREEN}✓  Passed  : {PASSED}{RESET}")
    print(f"  {RED}✗  Failed  : {FAILED}{RESET}")
    print(f"  {YELLOW}⚠  Warnings: {WARNINGS}{RESET}")
    if FAILED == 0:
        print(f"\n  {BOLD}{GREEN}✅ ALL CHECKS PASSED — Stage 7 ready for Demo sign-off{RESET}")
    else:
        print(f"\n  {BOLD}{RED}❌ {FAILED} check(s) failed{RESET}")
    print(f"{'═' * 65}\n")
    return 0 if FAILED == 0 else 1

def test_stage7_dashboard_suite() -> None:
    assert main() == 0


if __name__ == "__main__":
    sys.exit(main())
